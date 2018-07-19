import openpyxl
import calendar
import datetime


file=openpyxl.load_workbook("SMI.xlsx")
fs= file.get_sheet_by_name("Raw Data")

fs1=file.copy_worksheet(fs)
fs1.title="Processing"
fs1= file.get_sheet_by_name("Processing")
fs1.cell(row=1,column=4).value="Month"
fs1.cell(row=1,column=5).value="Year"
fs1.cell(row=1,column=6).value="No. of Weeks in Month"
fs1.cell(row=1,column=7).value="Facebook Impressions Per Week"
fs1.cell(row=1,column=8).value="Twitter Impressions Per Week"


def no_of_weeks(year,month):
    day_to_count = calendar.SATURDAY
    matrix = calendar.monthcalendar(year,month)
    return sum(1 for x in matrix if x[day_to_count] != 0)    

mlist=list(calendar.month_name)

print("Start Processing")
for row_id in range(1,fs1.max_row):
    
    for col_id in range(1,fs1.max_column):
        if(col_id==1 and fs1.cell(row=row_id,column=1).value!=None ):
            if(fs1.cell(row=row_id,column=2).value==None and fs1.cell(row=row_id,column=3).value==None):
                year=fs1.cell(row=row_id,column=col_id).value
            else:
                if(fs1.cell(row=row_id,column=1).value in calendar.month_name):
                    mnum=(mlist.index(fs1.cell(row=row_id,column=1).value))        
                    nweek=no_of_weeks(year,mnum)
                    fs1.cell(row=row_id,column=4).value=mnum
                    fs1.cell(row=row_id,column=6).value=nweek
                    fs1.cell(row=row_id,column=5).value=year

                    fb_imp=fs1.cell(row=row_id,column=2).value
                    twi_imp=fs1.cell(row=row_id,column=3).value
                    
                    fb_imp_week=fb_imp/nweek
                    fs1.cell(row=row_id,column=7).value=fb_imp_week
                    
                    twi_imp_week=twi_imp/nweek
                    fs1.cell(row=row_id,column=8).value=twi_imp_week    

print("Processing Done")

fs2=file.create_sheet("Final")
fs2= file.get_sheet_by_name("Final")
fs2.cell(row=1,column=1).value="Observation Week"
fs2.cell(row=1,column=2).value="Facebook"
fs2.cell(row=1,column=3).value="Twitter"
fs2.cell(row=1,column=4).value="Social Media Total"
fs2.cell(row=1,column=5).value="Month"
fs2.cell(row=1,column=6).value="Year"

week_sum=0
for col in fs1.iter_cols(min_row=2,min_col=6,max_row=fs1.max_row,max_col=6):
    for cell in col:      
        if(cell.value!=None):
            week_sum=week_sum+cell.value
            
print("Calculating Observation Week")            
x=datetime.date(2014,12,6)
row_i=2
for w in range(0,week_sum):
    fs2.cell(row=row_i,column=1).value=x
    row_i=row_i+1
    x=x+datetime.timedelta(days=7)
    
print("Preparing Final Sheet")
q=2
for row_id in range(2,fs1.max_row):
    nweek=fs1.cell(row=row_id,column=6).value
    if(nweek==None):
        continue
    year=fs1.cell(row=row_id,column=5).value
    month =fs1.cell(row=row_id,column=4).value
    fa_imp=fs1.cell(row=row_id,column=7).value
    twi_imp=fs1.cell(row=row_id,column=8).value
    l=1
    
    while (l<=nweek):
        fs2.cell(row=q,column=2).value=fa_imp
        fs2.cell(row=q,column=3).value=twi_imp
        fs2.cell(row=q,column=4).value=fa_imp+twi_imp
        fs2.cell(row=q,column=5).value=month
        fs2.cell(row=q,column=6).value=year
        l=l+1
        q=q+1
        
print("Final Sheet Done")

file.save('example_filetest.xlsx')
